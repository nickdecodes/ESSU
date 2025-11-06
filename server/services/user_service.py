# !/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author  : nickdecodes
@Email   : 
@Usage   : 
@Filename: user_service.py
@DateTime: 2025/10/25 23:17
@Software: vscode
"""

import logging
import time
import uuid
import json
import os
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from flask import send_file, jsonify

from dbs.db_manager import DBManager
from dbs.models import User, OperationRecord
from config import Config


class UserService:
    """用户服务 - 负责用户认证和用户管理"""
    
    def __init__(self):
        self.db = DBManager()
        self.logger = logging.getLogger(__name__)
    
    def _get_active_sessions(self, sessions_json: str) -> dict:
        """获取活跃会话"""
        sessions = json.loads(sessions_json or '{}')
        current_time = time.time()
        return {sid: ts for sid, ts in sessions.items() 
                if current_time - ts <= Config.SESSION_TIMEOUT}
    
    def _process_avatar_image(self, img_data, username: str) -> str:
        """处理头像图片:裁剪为正方形并压缩到200x200"""
        try:
            img_pil = Image.open(BytesIO(img_data))
            width, height = img_pil.size
            
            # 裁剪为中心正方形
            size = min(width, height)
            left = (width - size) // 2
            top = (height - size) // 2
            img_pil = img_pil.crop((left, top, left + size, top + size))
            
            # 压缩到200x200
            img_pil = img_pil.resize((200, 200), Image.LANCZOS)
            
            # 保存文件
            file_ext = 'png'
            unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
            file_path = os.path.join(Config.UPLOAD_FOLDER, unique_filename)
            img_pil.save(file_path)
            
            return f"images/{unique_filename}"
        except Exception as e:
            self.logger.warning(f'处理头像失败: {username} - {str(e)}')
            return None
    
    def _delete_avatar_file(self, avatar_path: str):
        """删除头像文件"""
        if not avatar_path or avatar_path.startswith('http'):
            return
        
        try:
            file_path = os.path.join(Config.UPLOAD_FOLDER, os.path.basename(avatar_path))
            if os.path.exists(file_path):
                os.remove(file_path)
                self.logger.info(f'删除头像成功: {avatar_path}')
        except Exception as e:
            self.logger.warning(f'删除头像失败: {avatar_path} - {str(e)}')
    
    def login(self, username: str, password: str):
        """用户登录"""
        self.logger.info(f'用户登录尝试: {username}')
        
        session = self.db.get_session()
        try:
            user = session.query(User).filter(
                User.username == username,
                User.password == password
            ).first()
            
            if not user:
                self.logger.warning(f'用户登录失败: {username} - 用户名或密码错误')
                return {'success': False, 'message': '用户名或密码错误'}
            
            # 清理过期会话
            self._cleanup_expired_sessions(user)
            
            # 检查并发登录数量（admin用户名无限制）
            sessions = json.loads(user.sessions or '{}')
            if username != 'admin' and len(sessions) >= Config.MAX_CONCURRENT_SESSIONS:
                self.logger.warning(f'用户登录失败: {username} - 超过最大并发登录数量')
                return {'success': False, 'message': f'同一用户最多只能同时在{Config.MAX_CONCURRENT_SESSIONS}个设备上登录'}
            
            # 创建新会话
            session_id = str(uuid.uuid4())
            sessions[session_id] = time.time()
            user.sessions = json.dumps(sessions)
            session.commit()
            
            self.logger.info(f'用户登录成功: {username}, 角色: {user.role}, 会话: {session_id}')
            return {
                'success': True, 
                'user': {
                    'username': user.username,
                    'role': user.role,
                    'avatar_path': user.avatar_path,
                    'session_id': session_id
                }
            }
        except Exception as e:
            self.logger.error(f'用户登录异常: {username} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '登录失败'}
        finally:
            self.db.close_session(session)

    def logout(self, username: str, session_id: str = None) -> bool:
        """用户登出"""
        session = self.db.get_session()
        try:
            user = session.query(User).filter(User.username == username).first()
            if user and session_id:
                sessions = json.loads(user.sessions or '{}')
                if session_id in sessions:
                    del sessions[session_id]
                    user.sessions = json.dumps(sessions)
                    self.logger.info(f'用户登出: {username}, 会话: {session_id}')
            
            # 记录登出操作
            operation = OperationRecord(
                operation_type='用户登出',
                name=username,
                quantity=0,
                detail=f'用户登出: {username}',
                username=username
            )
            session.add(operation)
            
            return self.db.commit_session(session)
        except Exception as e:
            session.rollback()
            self.logger.error(f'用户登出异常: {username} - {str(e)}', exc_info=True)
            return False
        finally:
            self.db.close_session(session)
    
    def add_user(self, username: str, password: str, role: str, operator: str = '', avatar_path: str = None, detail: str = None) -> bool:
        """添加用户"""
        self.logger.info(f'开始添加用户: {username}, 角色: {role}, 操作者: {operator}')
        
        session = self.db.get_session()
        try:
            # 创建用户对象
            user = User(
                username=username,
                password=password,
                role=role,
                avatar_path=avatar_path
            )
            
            session.add(user)
            session.flush()  # 获取ID但不提交
            
            # 记录添加操作
            operation = OperationRecord(
                operation_type='添加用户',
                name=username,
                quantity=0,
                detail=detail or f'添加用户: {username}, 角色: {role}',
                username=operator
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                self.logger.info(f'用户添加成功: {username} (ID: {user.id}), 角色: {role}')
                return True
            else:
                return False
                
        except Exception as e:
            session.rollback()
            if 'UNIQUE constraint failed' in str(e):
                self.logger.warning(f'用户添加失败: 用户名已存在 - {username}')
            else:
                self.logger.error(f'用户添加异常: {username} - {str(e)}', exc_info=True)
            return False
        finally:
            self.db.close_session(session)
    
    def update_user(self, user_id: int, username: str = None, password: str = None, role: str = None, operator: str = '', avatar_path: str = None, detail: str = None) -> dict:
        """更新用户信息"""
        session = self.db.get_session()
        try:
            user = session.query(User).filter(User.id == user_id).first()
            if not user:
                return {'success': False, 'message': '用户不存在'}
            
            # 检查用户名是否重复
            if username is not None and username != user.username:
                existing = session.query(User).filter(User.username == username, User.id != user_id).first()
                if existing:
                    return {'success': False, 'message': '用户名已存在'}
            
            old_avatar_path = user.avatar_path
            old_username = user.username
            
            # 更新字段
            updated_fields = []
            if username is not None and username != user.username:
                user.username = username
                updated_fields.append(f'修改用户名: {username}')
            if password is not None:
                user.password = password
                updated_fields.append('修改密码')
            if role is not None:
                user.role = role
                updated_fields.append(f'修改角色: {role}')
            if avatar_path is not None:
                user.avatar_path = avatar_path
                updated_fields.append('修改头像')
            
            if not updated_fields:
                return {'success': False, 'message': '没有需要更新的字段'}
            
            # 记录编辑操作
            operation = OperationRecord(
                operation_type='编辑用户',
                name=old_username,
                quantity=0,
                detail=detail or f'编辑用户: {old_username}, {"、".join(updated_fields)}',
                username=operator
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                # 删除旧头像文件
                if avatar_path is not None and old_avatar_path and old_avatar_path != avatar_path:
                    self._delete_avatar_file(old_avatar_path)
                return {'success': True}
            return {'success': False, 'message': '数据库操作失败'}
            
        except Exception as e:
            session.rollback()
            self.logger.error(f'用户更新异常: {user_id} - {str(e)}', exc_info=True)
            if 'UNIQUE constraint failed' in str(e):
                return {'success': False, 'message': '用户名已存在'}
            return {'success': False, 'message': '更新失败'}
        finally:
            self.db.close_session(session)
    
    def delete_user(self, username: str, operator: str = '') -> bool:
        """删除用户"""
        session = self.db.get_session()
        try:
            user = session.query(User).filter(User.username == username).first()
            if not user:
                return False
            
            avatar_path = user.avatar_path
            
            # 删除用户
            session.delete(user)
            
            # 记录删除操作
            operation = OperationRecord(
                operation_type='删除用户',
                name=username,
                quantity=0,
                detail=f'删除用户: {username}, 角色: {user.role}',
                username=operator
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                self._delete_avatar_file(avatar_path)
                return True
            return False
            
        except Exception as e:
            session.rollback()
            self.logger.error(f'用户删除异常: {username} - {str(e)}', exc_info=True)
            return False
        finally:
            self.db.close_session(session)
    
    def get_all_users(self):
        """获取所有用户"""
        session = self.db.get_session()
        try:
            users = session.query(User).order_by(User.id).all()
            result = []
            
            for u in users:
                active_sessions = self._get_active_sessions(u.sessions)
                
                result.append({
                    'id': u.id,
                    'username': u.username,
                    'password': u.password,
                    'role': u.role,
                    'avatar_path': u.avatar_path,
                    'online_devices': len(active_sessions),
                    'sessions': [{
                        'session_id': sid,
                        'login_time': ts,
                        'device_info': f'设备{i+1}'
                    } for i, (sid, ts) in enumerate(active_sessions.items())]
                })
            return result
        finally:
            self.db.close_session(session)
    
    def _cleanup_expired_sessions(self, user):
        """清理过期会话"""
        active_sessions = self._get_active_sessions(user.sessions)
        user.sessions = json.dumps(active_sessions)
    
    def remove_user_session(self, username: str, session_id: str, operator: str = '') -> dict:
        """移除用户指定会话"""
        session = self.db.get_session()
        try:
            user = session.query(User).filter(User.username == username).first()
            if not user:
                return {'success': False, 'message': '用户不存在'}
            
            sessions = json.loads(user.sessions or '{}')
            if session_id in sessions:
                del sessions[session_id]
                user.sessions = json.dumps(sessions)
                
                # 记录操作
                operation = OperationRecord(
                    operation_type='移除设备',
                    name=username,
                    quantity=0,
                    detail=f'管理员移除用户 {username} 的设备会话',
                    username=operator
                )
                session.add(operation)
                
                if self.db.commit_session(session):
                    self.logger.info(f'移除用户会话成功: {username}, 会话: {session_id}')
                    return {
                        'success': True, 
                        'removed_user': username,
                        'removed_session': session_id
                    }
                else:
                    return {'success': False, 'message': '数据库操作失败'}
            return {'success': False, 'message': '会话不存在'}
        except Exception as e:
            session.rollback()
            self.logger.error(f'移除用户会话异常: {username} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '操作失败'}
        finally:
            self.db.close_session(session)
    
    def import_from_excel(self, file, operator: str = '') -> dict:
        """从Excel导入用户（不支持头像导入）"""
        try:
            file_content = file.read()
            wb = openpyxl.load_workbook(BytesIO(file_content))
            ws = wb.active
            
            # 第一阶段：验证所有数据
            errors = []
            valid_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if not row[0].value:
                    continue
                
                try:
                    username = str(row[0].value).strip()
                    if not username:
                        errors.append(f'第{row_idx}行：用户名不能为空')
                        continue
                    
                    password = str(row[1].value).strip() if row[1].value else 'password'
                    role = str(row[2].value).strip() if row[2].value else 'user'
                    
                    if role not in Config.VALID_ROLES:
                        errors.append(f'第{row_idx}行（{username}）：角色"{role}"无效，有效值为{Config.VALID_ROLES}')
                        continue
                    
                    valid_rows.append((row_idx, row, username, password, role))
                    
                except Exception as e:
                    errors.append(f'第{row_idx}行：数据格式错误 - {str(e)}')
            
            # 如果有验证错误，直接返回
            if errors:
                return {
                    'success': False,
                    'message': '数据验证失败：\n' + '\n'.join(errors)
                }
            
            if not valid_rows:
                return {'success': False, 'message': '没有有效的数据行'}
            
            # 第二阶段：事务性执行导入（要么全部成功，要么全部失败）
            session = self.db.get_session()
            try:
                import_results = []
                
                # 预检查所有用户名是否会冲突
                for row_idx, row, username, password, role in valid_rows:
                    existing = session.query(User).filter(User.username == username).first()
                    import_results.append({
                        'username': username,
                        'password': password,
                        'role': role,
                        'existing': existing,
                        'action': 'update' if existing else 'create'
                    })
                
                # 执行所有用户操作
                for result in import_results:
                    if result['existing']:
                        # 更新现有用户
                        user = result['existing']
                        user.username = result['username']
                        user.password = result['password']
                        user.role = result['role']
                    else:
                        # 创建新用户
                        user = User(
                            username=result['username'],
                            password=result['password'],
                            role=result['role']
                        )
                        session.add(user)
                
                # 提交用户数据更改
                if self.db.commit_session(session):
                    success_count = len(valid_rows)
                    created_count = len([r for r in import_results if r['action'] == 'create'])
                    updated_count = len([r for r in import_results if r['action'] == 'update'])
                    
                    # 用户数据提交成功后，记录操作日志
                    log_session = self.db.get_session()
                    try:
                        for result in import_results:
                            detail = f'导入用户: {result["username"]}, 角色: {result["role"]}'
                            operation = OperationRecord(
                                operation_type='添加用户' if result['action'] == 'create' else '更新用户',
                                name=result['username'],
                                quantity=0,
                                detail=detail,
                                username=operator
                            )
                            log_session.add(operation)
                        self.db.commit_session(log_session)
                    finally:
                        self.db.close_session(log_session)
                    
                    return {
                        'success': True,
                        'total_count': success_count,
                        'created_count': created_count,
                        'updated_count': updated_count,
                        'message': f'导入成功！共处理 {success_count} 个用户（新增 {created_count} 个，更新 {updated_count} 个）\n注意：导入不支持头像，请手动上传用户头像'
                    }
                else:
                    return {'success': False, 'message': '数据库提交失败，所有操作已回滚'}
                    
            except Exception as e:
                session.rollback()
                self.logger.error(f'导入用户事务失败: {str(e)}', exc_info=True)
                return {'success': False, 'message': f'导入失败，所有操作已回滚：{str(e)}'}
            finally:
                self.db.close_session(session)
            
        except Exception as e:
            self.logger.error(f'导入Excel失败: {str(e)}', exc_info=True)
            return {'success': False, 'message': f'导入失败: {str(e)}'}
    
    def export_to_excel(self, user_id_list: list = None, operator: str = ''):
        """导出用户到Excel
        
        Args:
            user_id_list: 要导出的用户ID列表,如果为空则导出所有用户
            operator: 操作者用户名
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '用户列表'
            
            # 设置列宽和行高
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            
            # 写入表头
            headers = ['头像', '用户名', '密码', '角色', '创建时间', '更新时间']
            ws.append(headers)
            
            # 从数据库查询用户
            session = self.db.get_session()
            try:
                query = session.query(User).order_by(User.id)
                
                # 如果有筛选条件，只导出筛选后的用户
                if user_id_list:
                    query = query.filter(User.id.in_(user_id_list))
                
                users = query.all()
                
                row_idx = 2
                for user in users:
                    # 计算在线设备数
                    active_sessions = self._get_active_sessions(user.sessions)
                    
                    # 写入数据
                    ws.append([
                        '',
                        user.username,
                        user.password,
                        user.role,
                        user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
                        user.updated_at.strftime('%Y-%m-%d %H:%M:%S') if user.updated_at else ''
                    ])
                    
                    # 设置行高
                    ws.row_dimensions[row_idx].height = 50
                    
                    # 插入头像
                    if user.avatar_path:
                        try:
                            if user.avatar_path.startswith(('http://', 'https://')):
                                ws[f'A{row_idx}'] = user.avatar_path
                            else:
                                avatar_file = os.path.join(Config.UPLOAD_FOLDER, os.path.basename(user.avatar_path))
                                if os.path.exists(avatar_file):
                                    img = Image.open(avatar_file)
                                    img.thumbnail((50, 50))
                                    
                                    img_buffer = BytesIO()
                                    img.save(img_buffer, format='PNG')
                                    img_buffer.seek(0)
                                    
                                    xl_img = XLImage(img_buffer)
                                    xl_img.width = 50
                                    xl_img.height = 50
                                    ws.add_image(xl_img, f'A{row_idx}')
                        except Exception as e:
                            self.logger.warning(f'插入头像失败: {user.username} - {str(e)}')
                    
                    row_idx += 1
            finally:
                self.db.close_session(session)
            
            # 保存到内存
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 生成文件名
            filename = f'users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            # 记录导出操作
            if operator:
                session = self.db.get_session()
                try:
                    count = len(user_id_list) if user_id_list else len(users)
                    operation = OperationRecord(
                        operation_type='导出用户',
                        name=f'导出{count}个用户',
                        quantity=count,
                        detail=f'导出用户到Excel: {filename}',
                        username=operator
                    )
                    session.add(operation)
                    self.db.commit_session(session)
                finally:
                    self.db.close_session(session)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            self.logger.error(f'导出Excel失败: {str(e)}', exc_info=True)
            return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
    
    def generate_import_template(self):
        """生成用户导入模板"""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '用户导入模板'
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            
            # 写入表头
            headers = ['用户名', '密码', '角色']
            ws.append(headers)
            
            # 添加示例数据
            ws.append([
                'demo_user',
                'password123',
                'user'
            ])
            
            ws.append([
                'admin_user',
                'admin123',
                'admin'
            ])
            
            # 为角色列添加数据校验（下拉选择）
            dv = DataValidation(
                type="list",
                formula1='"admin,user"',
                allow_blank=True
            )
            dv.error = '请选择admin或user'
            dv.errorTitle = '角色错误'
            dv.prompt = '请选择用户角色'
            dv.promptTitle = '角色选择'
            
            # 应用到C列（角色列）的所有行
            ws.add_data_validation(dv)
            dv.add('C2:C1000')  # 从第2行开始到第1000行
            
            # 添加说明文本
            ws['A5'] = '导入说明：'
            ws['A6'] = '1. 用户名：必填，不能重复，最多50个字符'
            ws['A7'] = '2. 密码：可选，默认为password，最多100个字符'
            ws['A8'] = '3. 角色：可选，点击下拉选择admin或user，默认为user'
            ws['A9'] = '4. 头像：导入后请手动上传，不支持批量导入'
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='用户导入模板.xlsx'
            )
            
        except Exception as e:
            self.logger.error(f'生成导入模板失败: {str(e)}', exc_info=True)
            return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'}), 500
