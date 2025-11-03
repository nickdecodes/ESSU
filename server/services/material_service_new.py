#!/usr/bin/env python
# -*- coding: utf-8 -*-

from dbs.db_manager import DBManager
from dbs.models import Material, OperationRecord
from utils.timezone_utils import format_china_time
import logging

class MaterialService:
    """材料服务 - 负责材料的增删改查和库存管理"""
    
    def __init__(self):
        self.db = DBManager()
        self.logger = logging.getLogger(__name__)
    
    def add_material(self, name: str, in_price: float, out_price: float = None, username: str = '', image_path: str = None) -> dict:
        """添加材料"""
        if out_price is None:
            out_price = in_price
        
        session = self.db.get_session()
        try:
            self.logger.info(f'开始添加材料: {name}, 用户: {username}')
            
            material = Material(
                name=name,
                in_price=in_price,
                out_price=out_price,
                image_path=image_path
            )
            
            session.add(material)
            session.flush()
            
            operation = OperationRecord(
                operation_type='添加材料',
                name=name,
                quantity=0,
                detail=f'添加材料: {name}',
                username=username
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                self.logger.info(f'材料添加成功: {material.id} - {name}')
                return {'success': True, 'material_id': material.id}
            else:
                return {'success': False}
                
        except Exception as e:
            session.rollback()
            self.logger.error(f'材料添加异常: {name} - {str(e)}', exc_info=True)
            return {'success': False}
        finally:
            self.db.close_session(session)
    
    def get_all_materials(self):
        """获取所有材料"""
        session = self.db.get_session()
        try:
            materials = session.query(Material).order_by(Material.id).all()
            return [{
                'id': m.id,
                'name': m.name,
                'in_price': m.in_price,
                'out_price': m.out_price,
                'stock_count': m.stock_count,
                'image_path': m.image_path,
                'used_by_products': m.used_by_products,
                'created_at': format_china_time(m.created_at),
                'updated_at': format_china_time(m.updated_at)
            } for m in materials]
        finally:
            self.db.close_session(session)
    
    def get_materials_paginated(self, offset: int, limit: int):
        """分页获取材料"""
        session = self.db.get_session()
        try:
            materials = session.query(Material).order_by(Material.id).offset(offset).limit(limit).all()
            return [{
                'id': m.id,
                'name': m.name,
                'in_price': m.in_price,
                'out_price': m.out_price,
                'stock_count': m.stock_count,
                'image_path': m.image_path,
                'used_by_products': m.used_by_products,
                'created_at': format_china_time(m.created_at),
                'updated_at': format_china_time(m.updated_at)
            } for m in materials]
        finally:
            self.db.close_session(session)
    
    def get_materials_count(self) -> int:
        """获取材料总数"""
        session = self.db.get_session()
        try:
            return session.query(Material).count()
        finally:
            self.db.close_session(session)
    
    def _update_material_usage_count(self, session, material_id: int):
        """更新材料的使用计数"""
        from dbs.models import Product
        import json
        
        material = session.query(Material).filter(Material.id == material_id).first()
        if not material:
            return
        
        count = 0
        products = session.query(Product).all()
        material_id_str = str(material_id)
        
        for product in products:
            try:
                product_materials = json.loads(product.materials)
                if material_id_str in product_materials:
                    count += 1
            except (json.JSONDecodeError, ValueError):
                continue
        
        material.used_by_products = count
    
    def delete_material(self, material_id: int, username: str = '') -> dict:
        """删除材料"""
        self.logger.info(f'开始删除材料: {material_id}, 用户: {username}')
        
        session = self.db.get_session()
        try:
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                self.logger.warning(f'材料删除失败: 材料不存在 - {material_id}')
                return {'success': False, 'message': '材料不存在'}
            
            material_name = material.name
            stock_count = material.stock_count
            
            if stock_count > 0:
                return {'success': False, 'message': f'材料 {material_name} 库存不为零（{stock_count}个），请先出库后再删除'}
            
            if material.used_by_products > 0:
                return {'success': False, 'message': f'被{material.used_by_products}个产品使用，请先删除相关产品'}
            
            session.delete(material)
            
            operation = OperationRecord(
                operation_type='删除材料',
                name=material_name,
                quantity=0,
                detail=f'删除材料: {material_name}',
                username=username
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                self.logger.info(f'材料删除成功: {material_id} - {material_name}')
                return {'success': True}
            else:
                return {'success': False, 'message': '删除失败'}
                
        except Exception as e:
            session.rollback()
            self.logger.error(f'材料删除异常: {material_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '删除失败'}
        finally:
            self.db.close_session(session)
    
    def batch_delete_materials(self, material_ids: list, username: str = '') -> dict:
        """批量删除材料"""
        try:
            failed_materials = []
            deleted_count = 0
            
            for material_id in material_ids:
                result = self.delete_material(material_id, username)
                if result['success']:
                    deleted_count += 1
                else:
                    failed_materials.append({
                        'id': material_id,
                        'message': result['message']
                    })
            
            if failed_materials:
                failed_list = []
                for failed in failed_materials:
                    session = self.db.get_session()
                    try:
                        material = session.query(Material).filter(Material.id == failed['id']).first()
                        if material:
                            failed_list.append({
                                'name': material.name,
                                'reason': failed['message']
                            })
                    finally:
                        self.db.close_session(session)
                
                return {
                    'success': False,
                    'deleted_count': deleted_count,
                    'failed_materials': failed_list,
                    'should_close': deleted_count > 0
                }
            else:
                return {'success': True, 'message': f'成功删除 {deleted_count} 个材料'}
        except Exception as e:
            return {'success': False, 'message': '批量删除失败'}
    
    def check_related_products(self, material_id: int, in_price: float = None, out_price: float = None) -> dict:
        """检查使用该材料的产品列表"""
        session = self.db.get_session()
        try:
            from dbs.models import Product
            import json
            
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                return {'success': False, 'message': '材料不存在'}
            
            price_changed = False
            if in_price is not None and material.in_price != in_price:
                price_changed = True
            if out_price is not None and material.out_price != out_price:
                price_changed = True
            
            if not price_changed:
                return {'success': True, 'price_changed': False, 'affected_products': []}
            
            products = session.query(Product).all()
            material_id_str = str(material_id)
            affected_products = []
            
            for product in products:
                try:
                    materials = json.loads(product.materials)
                    if material_id_str in materials:
                        current_cost = 0
                        current_selling = 0
                        new_cost = 0
                        new_selling = 0
                        
                        for mat_id, required_qty in materials.items():
                            mat = session.query(Material).filter(Material.id == int(mat_id)).first()
                            if mat:
                                current_cost += (mat.in_price or 0) * required_qty
                                current_selling += (mat.out_price or 0) * required_qty
                                
                                if int(mat_id) == material_id:
                                    new_cost += (in_price if in_price is not None else mat.in_price or 0) * required_qty
                                    new_selling += (out_price if out_price is not None else mat.out_price or 0) * required_qty
                                else:
                                    new_cost += (mat.in_price or 0) * required_qty
                                    new_selling += (mat.out_price or 0) * required_qty
                        
                        manual_price = product.manual_price or 0
                        current_selling += manual_price
                        new_selling += manual_price
                        
                        material_list = []
                        for mat_id, qty in materials.items():
                            mat = session.query(Material).filter(Material.id == int(mat_id)).first()
                            if mat:
                                material_list.append(f'{mat.name}×{qty}')
                        
                        affected_products.append({
                            'id': product.id,
                            'name': product.name,
                            'image_path': product.image_path,
                            'materials': ', '.join(material_list),
                            'current_cost': round(current_cost, 2),
                            'current_selling': round(current_selling, 2),
                            'new_cost': round(new_cost, 2),
                            'new_selling': round(new_selling, 2)
                        })
                except (json.JSONDecodeError, ValueError):
                    continue
            
            return {
                'success': True,
                'price_changed': True,
                'affected_products': affected_products
            }
            
        except Exception as e:
            self.logger.error(f'检查相关产品失败: {str(e)}', exc_info=True)
            return {'success': False, 'message': '检查失败'}
        finally:
            self.db.close_session(session)
    
    def update_material(self, material_id: int, name: str, in_price: float = None, out_price: float = None, username: str = '', image_path: str = None) -> bool:
        """更新材料信息"""
        session = self.db.get_session()
        try:
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                return False
            
            price_changed = False
            if in_price is not None and material.in_price != in_price:
                price_changed = True
            if out_price is not None and material.out_price != out_price:
                price_changed = True
            
            material.name = name
            if in_price is not None:
                material.in_price = in_price
            if out_price is not None:
                material.out_price = out_price
            if image_path is not None:
                material.image_path = image_path
            
            if price_changed:
                self._update_related_products_price(session, material_id)
            
            operation = OperationRecord(
                operation_type='更新材料',
                name=name,
                quantity=0,
                detail=f'更新材料: {name}',
                username=username
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                self.logger.info(f'材料更新成功: {material_id} - {name}')
                return True
            else:
                return False
                
        except Exception as e:
            session.rollback()
            self.logger.error(f'材料更新异常: {material_id} - {str(e)}', exc_info=True)
            return False
        finally:
            self.db.close_session(session)
    
    def _update_related_products_price(self, session, material_id: int):
        """更新使用该材料的产品价格"""
        from dbs.models import Product
        import json
        
        try:
            products = session.query(Product).all()
            material_id_str = str(material_id)
            
            for product in products:
                try:
                    materials = json.loads(product.materials)
                    
                    if material_id_str in materials:
                        total_in_price = 0
                        total_out_price = 0
                        
                        for mat_id, required_qty in materials.items():
                            material = session.query(Material).filter(Material.id == int(mat_id)).first()
                            if material:
                                total_in_price += (material.in_price or 0) * required_qty
                                total_out_price += (material.out_price or 0) * required_qty
                        
                        product.in_price = total_in_price
                        product.out_price = total_out_price + (product.manual_price or 0)
                        
                        self.logger.info(f'更新产品价格: {product.name} - 成本: {total_in_price}, 售价: {product.out_price}')
                        
                except (json.JSONDecodeError, ValueError) as e:
                    self.logger.warning(f'解析产品材料失败: {product.id} - {str(e)}')
                    continue
                    
        except Exception as e:
            self.logger.error(f'更新相关产品价格失败: {str(e)}', exc_info=True)
    
    def inbound(self, material_id: int, quantity: int, supplier: str, username: str = '') -> bool:
        """入库材料"""
        self.logger.info(f'开始材料入库: {material_id}, 数量: {quantity}, 用户: {username}')
        
        session = self.db.get_session()
        try:
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                self.logger.warning(f'材料入库失败: 材料不存在 - {material_id}')
                return False
            
            material.stock_count += quantity
            
            operation = OperationRecord(
                operation_type='材料入库',
                name=material.name,
                quantity=quantity,
                detail=f'供应商: {supplier}, 数量: +{quantity}',
                username=username
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                self.logger.info(f'材料入库成功: {material_id}, 数量: {quantity}')
                return True
            else:
                self.logger.error(f'材料入库失败: 数据库更新失败 - {material_id}')
                return False
                
        except Exception as e:
            session.rollback()
            self.logger.error(f'材料入库异常: {material_id} - {str(e)}', exc_info=True)
            return False
        finally:
            self.db.close_session(session)
    
    def outbound(self, material_id: int, quantity: int, customer: str, username: str = '') -> dict:
        """出库操作"""
        session = self.db.get_session()
        try:
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                return {'success': False, 'message': '材料不存在'}
            
            current_stock = material.stock_count
            if current_stock < quantity:
                return {'success': False, 'message': f'库存不足，当前库存: {current_stock}'}
            
            material.stock_count -= quantity
            
            operation = OperationRecord(
                operation_type='材料出库',
                name=material.name,
                quantity=-quantity,
                detail=f'客户: {customer}, 数量: -{quantity}',
                username=username
            )
            session.add(operation)
            
            if self.db.commit_session(session):
                return {'success': True}
            else:
                return {'success': False, 'message': '出库失败'}
                
        except Exception as e:
            session.rollback()
            self.logger.error(f'材料出库异常: {material_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '出库失败'}
        finally:
            self.db.close_session(session)
    
    def get_stock(self, material_id: int) -> int:
        """获取材料库存"""
        session = self.db.get_session()
        try:
            material = session.query(Material).filter(Material.id == material_id).first()
            return material.stock_count if material else 0
        finally:
            self.db.close_session(session)
    
    def import_from_excel(self, file, username: str = '') -> dict:
        """从Excel导入材料"""
        import openpyxl
        from io import BytesIO
        
        try:
            self.logger.info(f'开始导入Excel, 用户: {username}')
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            self.logger.info(f'Excel读取成功, 工作表: {ws.title}')
            
            success_count = 0
            failed_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                first_col = str(row[0]) if row[0] else ''
                is_new_format = len(row) >= 7 and (not row[0] or first_col.startswith(('http', 'images/', '#')))
                
                if is_new_format:
                    name_idx, price_in_idx, price_out_idx, stock_idx = 1, 2, 3, 4
                else:
                    name_idx, price_in_idx, price_out_idx, stock_idx = 0, 1, 2, 3
                
                if not row[name_idx]:
                    continue
                
                try:
                    name = str(row[name_idx]).strip()
                    in_price = float(row[price_in_idx]) if row[price_in_idx] else 0
                    out_price = float(row[price_out_idx]) if row[price_out_idx] else in_price
                    stock_count = int(row[stock_idx]) if stock_idx < len(row) and row[stock_idx] else 0
                    
                    session = self.db.get_session()
                    existing = session.query(Material).filter(Material.name == name).first()
                    
                    if existing:
                        existing.in_price = in_price
                        existing.out_price = out_price
                        existing.stock_count = stock_count
                        
                        operation = OperationRecord(
                            operation_type='导入材料',
                            name=name,
                            quantity=stock_count,
                            detail=f'导入材料: {name}, 库存: {stock_count}',
                            username=username
                        )
                        session.add(operation)
                        
                        if self.db.commit_session(session):
                            success_count += 1
                        else:
                            failed_count += 1
                    else:
                        material = Material(
                            name=name,
                            in_price=in_price,
                            out_price=out_price,
                            stock_count=stock_count
                        )
                        session.add(material)
                        session.flush()
                        
                        operation = OperationRecord(
                            operation_type='导入材料',
                            name=name,
                            quantity=stock_count,
                            detail=f'导入材料: {name}, 库存: {stock_count}',
                            username=username
                        )
                        session.add(operation)
                        
                        if self.db.commit_session(session):
                            success_count += 1
                        else:
                            failed_count += 1
                    
                    self.db.close_session(session)
                except Exception as e:
                    session.rollback()
                    self.db.close_session(session)
                    self.logger.error(f'第{row_num}行导入失败: {str(e)}', exc_info=True)
                    failed_count += 1
            
            msg = f'导入完成，成功 {success_count} 个，失败 {failed_count} 个'
            self.logger.info(msg)
            return {'success': True, 'message': msg}
            
        except Exception as e:
            msg = f'导入Excel失败: {str(e)}'
            self.logger.error(msg, exc_info=True)
            return {'success': False, 'message': msg}
    
    def export_to_excel(self, material_ids: list = None):
        """导出材料到Excel"""
        import openpyxl
        from flask import send_file
        from io import BytesIO
        from datetime import datetime
        from PIL import Image
        from openpyxl.drawing.image import Image as XLImage
        from config import Config
        import os
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '材料列表'
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            
            headers = ['图片', '材料名称', '进价', '售价', '库存数量', '创建时间', '更新时间']
            ws.append(headers)
            
            session = self.db.get_session()
            try:
                query = session.query(Material).order_by(Material.id)
                
                if material_ids:
                    query = query.filter(Material.id.in_(material_ids))
                
                materials = query.all()
                
                row_idx = 2
                for material in materials:
                    ws.append([
                        '',
                        material.name,
                        material.in_price,
                        material.out_price,
                        material.stock_count,
                        format_china_time(material.created_at),
                        format_china_time(material.updated_at)
                    ])
                    
                    ws.row_dimensions[row_idx].height = 50
                    
                    if material.image_path:
                        try:
                            if material.image_path.startswith(('http://', 'https://')):
                                ws[f'A{row_idx}'] = material.image_path
                            else:
                                image_file = os.path.join(Config.UPLOAD_FOLDER, os.path.basename(material.image_path))
                                if os.path.exists(image_file):
                                    img = Image.open(image_file)
                                    img.thumbnail((50, 50))
                                    
                                    img_buffer = BytesIO()
                                    img.save(img_buffer, format='PNG')
                                    img_buffer.seek(0)
                                    
                                    xl_img = XLImage(img_buffer)
                                    xl_img.width = 50
                                    xl_img.height = 50
                                    ws.add_image(xl_img, f'A{row_idx}')
                        except Exception as e:
                            self.logger.warning(f'插入图片失败: {material.name} - {str(e)}')
                    
                    row_idx += 1
            finally:
                self.db.close_session(session)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'materials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            self.logger.error(f'导出Excel失败: {str(e)}', exc_info=True)
            from flask import jsonify
            return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
