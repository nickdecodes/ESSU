#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author  : nickdecodes
@Email   : 
@Usage   : 
@Filename: material_service.py
@DateTime: 2025/10/25 23:17
@Software: vscode
"""

import os
import json
import logging
import openpyxl
from io import BytesIO
from dbs.db_manager import DBManager
from datetime import datetime
from PIL import Image
from flask import jsonify, send_file
from openpyxl.drawing.image import Image as XLImage
from dbs.models import Material, Product
from utils.timezone_utils import format_china_time
from config import Config


class MaterialService:
    """材料服务 - 负责材料的增删改查和库存管理"""
    
    def __init__(self):
        self.db = DBManager()
        self.logger = logging.getLogger(__name__)
    
    def _parse_used_list(self, used_by_products_str: str) -> list:
        """解析used_by_products JSON字符串"""
        try:
            return json.loads(used_by_products_str or '[]')
        except (json.JSONDecodeError, TypeError):
            self.logger.warning(f'解析used_by_products失败: {used_by_products_str}')
            return []
    
    def get_products_using_material(self, material_id: int) -> list:
        """获取使用该材料的产品ID列表"""
        with self.db.session_scope() as session:
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                return []
            return self._parse_used_list(material.used_by_products)
    
    def add_material(self, name: str, in_price: float, out_price: float = None, image_path: str = None, stock_count: int = 0) -> dict:
        """添加材料"""
        if out_price is None:
            out_price = in_price
        
        try:
            with self.db.session_scope() as session:
                material = Material(
                    name=name,
                    in_price=in_price,
                    out_price=out_price,
                    image_path=image_path,
                    stock_count=stock_count
                )
                session.add(material)
                session.flush()
                self.logger.info(f'材料添加成功: {material.id} - {name}')
                return {'success': True, 'material_id': material.id}
        except Exception as e:
            self.logger.error(f'材料添加异常: {name} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '添加失败'}
    
    def get_all_materials(self):
        """获取所有材料"""
        with self.db.session_scope() as session:
            materials = session.query(Material).order_by(Material.id).all()
            return {'success': True, 'materials': [{
                'id': m.id,
                'name': m.name,
                'in_price': m.in_price,
                'out_price': m.out_price,
                'stock_count': m.stock_count,
                'image_path': m.image_path,
                'used_by_products': self._parse_used_list(m.used_by_products),
                'is_used': len(self._parse_used_list(m.used_by_products)) > 0,
                'created_at': format_china_time(m.created_at),
                'updated_at': format_china_time(m.updated_at)
            } for m in materials]}
    
    def get_materials_paginated(self, offset: int, limit: int):
        """分页获取材料"""
        with self.db.session_scope() as session:
            materials = session.query(Material).order_by(Material.id).offset(offset).limit(limit).all()
            return {'success': True, 'materials': [{
                'id': m.id,
                'name': m.name,
                'in_price': m.in_price,
                'out_price': m.out_price,
                'stock_count': m.stock_count,
                'image_path': m.image_path,
                'used_by_products': self._parse_used_list(m.used_by_products),
                'is_used': len(self._parse_used_list(m.used_by_products)) > 0,
                'created_at': format_china_time(m.created_at),
                'updated_at': format_china_time(m.updated_at)
            } for m in materials]}
    
    def get_materials_count(self) -> dict:
        """获取材料总数"""
        with self.db.session_scope() as session:
            return {'success': True, 'count': session.query(Material).count()}
    
    def delete_material(self, material_id: int) -> dict:
        """删除材料"""
        try:
            with self.db.session_scope() as session:
                material = session.query(Material).filter(Material.id == material_id).first()
                if not material:
                    return {'success': False, 'message': '材料不存在'}
                
                if material.stock_count > 0:
                    return {'success': False, 'message': f'材料 {material.name} 库存不为零（{material.stock_count}个），请先出库后再删除'}
                
                used_list = self._parse_used_list(material.used_by_products)
                if len(used_list) > 0:
                    return {'success': False, 'message': f'被{len(used_list)}个产品使用，请先删除相关产品'}
                
                session.delete(material)
                self.logger.info(f'材料删除成功: {material_id} - {material.name}')
                return {'success': True}
        except Exception as e:
            self.logger.error(f'材料删除异常: {material_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '删除失败'}
    
    def batch_delete_materials(self, material_ids: list) -> dict:
        """批量删除材料"""
        try:
            with self.db.session_scope() as session:
                materials = session.query(Material).filter(Material.id.in_(material_ids)).all()
                failed_materials = []
                deleted_count = 0
                
                for material in materials:
                    if material.stock_count > 0:
                        failed_materials.append({
                            'name': material.name,
                            'reason': f'库存不为零（{material.stock_count}个），请先出库后再删除'
                        })
                        continue
                    
                    used_list = self._parse_used_list(material.used_by_products)
                    if len(used_list) > 0:
                        failed_materials.append({
                            'name': material.name,
                            'reason': f'被{len(used_list)}个产品使用，请先删除相关产品'
                        })
                        continue
                    
                    session.delete(material)
                    deleted_count += 1
                
                self.logger.info(f'批量删除材料: 成功{deleted_count}个')
                if failed_materials:
                    return {
                        'success': False,
                        'deleted_count': deleted_count,
                        'failed_materials': failed_materials,
                        'should_close': deleted_count > 0
                    }
                return {'success': True, 'message': f'成功删除 {deleted_count} 个材料'}
        except Exception as e:
            self.logger.error(f'批量删除材料异常: {str(e)}', exc_info=True)
            return {'success': False, 'message': '删除失败'}
    
    def check_related_products(self, material_id: int, in_price: float = None, out_price: float = None) -> dict:
        """检查使用该材料的产品列表"""
        try:
            with self.db.session_scope() as session:
                material = session.query(Material).filter(Material.id == material_id).first()
                if not material:
                    return {'success': False, 'message': '材料不存在'}
                
                price_changed = False
                if in_price is not None and material.in_price != in_price:
                    price_changed = True
                if out_price is not None and material.out_price != out_price:
                    price_changed = True
                
                if not price_changed:
                    return {'success': True, 'price_changed': False, 'affected_products': []}
                
                products = session.query(Product).all()
                material_id_str = str(material_id)
                affected_products = []
                
                for product in products:
                    try:
                        materials = json.loads(product.materials)
                        if material_id_str in materials:
                            current_cost = 0
                            current_selling = 0
                            new_cost = 0
                            new_selling = 0
                            
                            for mat_id, required_qty in materials.items():
                                mat = session.query(Material).filter(Material.id == int(mat_id)).first()
                                if mat:
                                    current_cost += (mat.in_price or 0) * required_qty
                                    current_selling += (mat.out_price or 0) * required_qty
                                    
                                    if int(mat_id) == material_id:
                                        new_cost += (in_price if in_price is not None else mat.in_price or 0) * required_qty
                                        new_selling += (out_price if out_price is not None else mat.out_price or 0) * required_qty
                                    else:
                                        new_cost += (mat.in_price or 0) * required_qty
                                        new_selling += (mat.out_price or 0) * required_qty
                            
                            other_price = product.other_price or 0
                            current_selling += other_price
                            new_selling += other_price
                            
                            material_list = []
                            for mat_id, qty in materials.items():
                                mat = session.query(Material).filter(Material.id == int(mat_id)).first()
                                if mat:
                                    material_list.append(f'{mat.name}×{qty}')
                            
                            affected_products.append({
                                'id': product.id,
                                'name': product.name,
                                'image_path': product.image_path,
                                'materials': ', '.join(material_list),
                                'current_cost': round(current_cost, 2),
                                'current_selling': round(current_selling, 2),
                                'new_cost': round(new_cost, 2),
                                'new_selling': round(new_selling, 2)
                            })
                    except (json.JSONDecodeError, ValueError):
                        continue
                
                return {
                    'success': True,
                    'price_changed': True,
                    'affected_products': affected_products
                }
            
        except Exception as e:
            self.logger.error(f'检查相关产品失败: {str(e)}', exc_info=True)
            return {'success': False, 'message': '检查失败'}
    
    def update_material(self, material_id: int, name: str, in_price: float = None, out_price: float = None, image_path: str = None, stock_count: int = None) -> dict:
        """更新材料信息"""
        try:
            with self.db.session_scope() as session:
                material = session.query(Material).filter(Material.id == material_id).first()
                if not material:
                    return {'success': False, 'message': '材料不存在'}
                
                price_changed = False
                if in_price is not None and material.in_price != in_price:
                    price_changed = True
                if out_price is not None and material.out_price != out_price:
                    price_changed = True
                
                material.name = name
                if in_price is not None:
                    material.in_price = in_price
                if out_price is not None:
                    material.out_price = out_price
                if image_path is not None:
                    material.image_path = image_path
                if stock_count is not None:
                    material.stock_count = stock_count
                
                if price_changed:
                    self._update_related_products_price(session, material_id)
                
                self.logger.info(f'材料更新成功: {material_id} - {name}')
                return {'success': True, 'price_changed': price_changed}
        except Exception as e:
            self.logger.error(f'材料更新异常: {material_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '更新失败'}
    
    def _update_related_products_price(self, session, material_id: int):
        """更新使用该材料的产品价格"""
        try:
            # 优化：一次性加载所有材料，避免N+1查询
            all_materials = session.query(Material).all()
            material_map = {str(m.id): m for m in all_materials}
            
            products = session.query(Product).all()
            material_id_str = str(material_id)
            
            for product in products:
                try:
                    materials = json.loads(product.materials)
                    
                    if material_id_str in materials:
                        total_in_price = 0
                        
                        for mat_id, required_qty in materials.items():
                            material = material_map.get(mat_id)
                            if material:
                                total_in_price += (material.in_price or 0) * required_qty
                        
                        product.in_price = total_in_price
                        product.out_price = total_in_price + (product.other_price or 0)
                        
                        self.logger.debug(f'更新产品价格: {product.name} - 成本: {total_in_price}, 售价: {product.out_price}')
                        
                except (json.JSONDecodeError, ValueError) as e:
                    self.logger.warning(f'解析产品材料失败: {product.id} - {str(e)}')
                    continue
                    
        except Exception as e:
            self.logger.error(f'更新相关产品价格失败: {str(e)}', exc_info=True)
    
    def inbound(self, material_id: int, quantity: int, supplier: str) -> dict:
        """入库材料"""
        try:
            with self.db.session_scope() as session:
                material = session.query(Material).filter(Material.id == material_id).first()
                if not material:
                    return {'success': False, 'message': '材料不存在'}
                
                material.stock_count += quantity
                self.logger.info(f'材料入库成功: {material_id}, 数量: {quantity}')
                return {'success': True, 'material_name': material.name}
        except Exception as e:
            self.logger.error(f'材料入库异常: {material_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '入库失败'}
    
    def outbound(self, material_id: int, quantity: int, customer: str) -> dict:
        """出库操作"""
        try:
            with self.db.session_scope() as session:
                material = session.query(Material).filter(Material.id == material_id).first()
                if not material:
                    return {'success': False, 'message': '材料不存在'}
                
                if material.stock_count < quantity:
                    return {'success': False, 'message': f'库存不足，当前库存: {material.stock_count}'}
                
                material.stock_count -= quantity
                self.logger.info(f'材料出库成功: {material_id}, 数量: {quantity}')
                return {'success': True, 'material_name': material.name}
        except Exception as e:
            self.logger.error(f'材料出库异常: {material_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '出库失败'}
    
    def get_stock(self, material_id: int) -> dict:
        """获取材料库存"""
        with self.db.session_scope() as session:
            material = session.query(Material).filter(Material.id == material_id).first()
            if not material:
                return {'success': False, 'message': '材料不存在'}
            return {'success': True, 'stock': material.stock_count or 0}
    
    def import_from_excel(self, file) -> dict:
        """从Excel导入材料"""
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            failed_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                first_col = str(row[0]) if row[0] else ''
                is_new_format = len(row) >= 7 and (not row[0] or first_col.startswith(('http', 'images/', '#')))
                
                if is_new_format:
                    name_idx, price_in_idx, price_out_idx, stock_idx = 1, 2, 3, 4
                else:
                    name_idx, price_in_idx, price_out_idx, stock_idx = 0, 1, 2, 3
                
                if not row[name_idx]:
                    continue
                
                try:
                    name = str(row[name_idx]).strip()
                    in_price = float(row[price_in_idx]) if row[price_in_idx] else 0
                    out_price = float(row[price_out_idx]) if row[price_out_idx] else in_price
                    import_stock = int(row[stock_idx]) if stock_idx < len(row) and row[stock_idx] else 0
                    
                    with self.db.session_scope() as session:
                        existing = session.query(Material).filter(Material.name == name).first()
                    
                    if existing:
                        # 检查价格是否匹配
                        if existing.in_price != in_price:
                            self.logger.error(f'第{row_num}行导入失败: {name} 进价不匹配，数据库: {existing.in_price}, 导入: {in_price}')
                            failed_count += 1
                            continue
                        if existing.out_price != out_price:
                            self.logger.error(f'第{row_num}行导入失败: {name} 售价不匹配，数据库: {existing.out_price}, 导入: {out_price}')
                            failed_count += 1
                            continue
                        
                        # 库存自增
                        new_stock = existing.stock_count + import_stock
                        result = self.update_material(
                            existing.id, name, in_price, out_price, None, new_stock
                        )
                        if result.get('success'):
                            updated_count += 1
                        else:
                            failed_count += 1
                    else:
                        result = self.add_material(
                            name, in_price, out_price, None, import_stock
                        )
                        if result.get('success'):
                            created_count += 1
                        else:
                            failed_count += 1
                            
                except Exception as e:
                    self.logger.error(f'第{row_num}行导入失败: {str(e)}', exc_info=True)
                    failed_count += 1
            
            msg = f'导入完成，成功 {created_count + updated_count} 个，失败 {failed_count} 个'
            return {
                'success': True, 
                'message': msg,
                'total_count': created_count + updated_count + failed_count,
                'created_count': created_count,
                'updated_count': updated_count
            }
            
        except Exception as e:
            msg = f'导入Excel失败: {str(e)}'
            self.logger.error(msg, exc_info=True)
            return {'success': False, 'message': msg}
    
    def export_to_excel(self, material_ids: list = None):
        """导出材料到Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '材料列表'
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            
            headers = ['图片', '材料名称', '进价', '售价', '库存数量', '创建时间', '更新时间']
            ws.append(headers)
            
            with self.db.session_scope() as session:
                query = session.query(Material).order_by(Material.id)
                
                if material_ids:
                    query = query.filter(Material.id.in_(material_ids))
                
                materials = query.all()
                
                row_idx = 2
                for material in materials:
                    ws.append([
                        '',
                        material.name,
                        material.in_price,
                        material.out_price,
                        material.stock_count,
                        format_china_time(material.created_at),
                        format_china_time(material.updated_at)
                    ])
                    
                    ws.row_dimensions[row_idx].height = 50
                    
                    if material.image_path:
                        try:
                            if material.image_path.startswith(('http://', 'https://')):
                                ws[f'A{row_idx}'] = material.image_path
                            else:
                                image_file = os.path.join(Config.UPLOAD_FOLDER, os.path.basename(material.image_path))
                                if os.path.exists(image_file):
                                    img = Image.open(image_file)
                                    img.thumbnail((50, 50))
                                    
                                    img_buffer = BytesIO()
                                    img.save(img_buffer, format='PNG')
                                    img_buffer.seek(0)
                                    
                                    xl_img = XLImage(img_buffer)
                                    xl_img.width = 50
                                    xl_img.height = 50
                                    ws.add_image(xl_img, f'A{row_idx}')
                        except Exception as e:
                            self.logger.warning(f'插入图片失败: {material.name} - {str(e)}')
                    
                    row_idx += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'materials_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            self.logger.error(f'导出Excel失败: {str(e)}', exc_info=True)
            return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
    
    def generate_import_template(self):
        """生成材料导入模板"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '材料导入模板'
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            
            # 写入表头
            headers = ['材料名称', '进价', '售价', '数量']
            ws.append(headers)
            
            # 添加示例数据
            ws.append([
                '红玛瑙珠子',
                10.50,
                15.00,
                100
            ])
            
            ws.append([
                '银线',
                2.30,
                3.50,
                50
            ])
            
            # 添加说明文本
            ws['A5'] = '导入说明：'
            ws['A6'] = '1. 材料名称：必填，不能重复，最多100个字符'
            ws['A7'] = '2. 进价：必填，数字类型'
            ws['A8'] = '3. 售价：可选，默认为进价，数字类型'
            ws['A9'] = '4. 数量：可选，默认为0，导入后会加到库存中'
            ws['A10'] = '5. 如果材料名称已存在，将更新该材料的信息'
            ws['A11'] = '6. 导入不支持图片，请在导入后手动上传材料图片'
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='材料导入模板.xlsx'
            )
            
        except Exception as e:
            self.logger.error(f'生成材料导入模板失败: {str(e)}', exc_info=True)
            return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'}), 500
