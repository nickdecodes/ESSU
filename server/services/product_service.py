#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author  : nickdecodes
@Email   : 
@Usage   : 
@Filename: product_service.py
@DateTime: 2025/10/25 23:17
@Software: vscode
"""

import os
import json
import logging
import openpyxl
from io import BytesIO
from flask import send_file, jsonify
from datetime import datetime
from PIL import Image
from openpyxl.drawing.image import Image as XLImage
from dbs.db_manager import DBManager
from dbs.models import Product, Material
from services.material_service import MaterialService
from utils.timezone_utils import format_china_time
from config import Config


class ProductService:
    """产品服务 - 负责产品配方管理和生产操作"""
    
    def __init__(self):
        self.db = DBManager()
        self.material_service = MaterialService()
        self.logger = logging.getLogger(__name__)
    
    def _parse_used_list(self, used_by_products_str: str) -> list:
        """解析used_by_products JSON字符串"""
        try:
            return json.loads(used_by_products_str or '[]')
        except (json.JSONDecodeError, TypeError):
            return []
    
    def _serialize_used_list(self, used_list: list) -> str:
        """序列化used_by_products列表"""
        return json.dumps(used_list)
    
    def _validate_materials(self, session, materials: dict) -> tuple:
        """验证材料是否存在，返回(是否有效, 错误信息, 材料对象字典)"""
        if not materials:
            return True, '', {}
        
        material_objs = {}
        for material_id_str in materials.keys():
            try:
                material_id = int(material_id_str)
                material = session.query(Material).filter(Material.id == material_id).first()
                if not material:
                    return False, f'材料ID {material_id} 不存在', {}
                material_objs[material_id_str] = material
            except (ValueError, TypeError):
                return False, f'无效的材料ID: {material_id_str}', {}
        
        return True, '', material_objs
    
    def _update_material_references(self, session, product_id: int, old_materials: dict, new_materials: dict):
        """批量更新材料引用"""
        try:
            old_ids = set(old_materials.keys())
            new_ids = set(new_materials.keys())
            
            for material_id_str in old_ids - new_ids:
                material = session.query(Material).filter(Material.id == int(material_id_str)).first()
                if material:
                    used_list = self._parse_used_list(material.used_by_products)
                    if product_id in used_list:
                        used_list.remove(product_id)
                        material.used_by_products = self._serialize_used_list(used_list)
            
            for material_id_str in new_ids - old_ids:
                material = session.query(Material).filter(Material.id == int(material_id_str)).first()
                if material:
                    used_list = self._parse_used_list(material.used_by_products)
                    if product_id not in used_list:
                        used_list.append(product_id)
                        material.used_by_products = self._serialize_used_list(used_list)
            
            return True
        except Exception as e:
            self.logger.error(f'更新材料引用失败: {str(e)}', exc_info=True)
            return False
    
    def add_product(self, name: str, materials: dict, in_price: float = 0, out_price: float = 0, other_price: float = 0, image_path: str = None) -> dict:
        """添加产品"""
        try:
            with self.db.session_scope() as session:
                valid, error_msg, material_objs = self._validate_materials(session, materials)
                if not valid:
                    return {'success': False, 'message': error_msg}
                
                if materials:
                    calculated_in_price = sum((material_objs.get(mid).in_price or 0) * qty for mid, qty in materials.items())
                    final_in_price = calculated_in_price
                    final_out_price = final_in_price + (other_price or 0)
                else:
                    final_in_price = in_price
                    final_out_price = out_price or final_in_price + (other_price or 0)
                
                product = Product(
                    name=name,
                    materials=json.dumps(materials),
                    in_price=final_in_price,
                    out_price=final_out_price,
                    other_price=other_price or 0,
                    image_path=image_path
                )
                
                session.add(product)
                session.flush()
                
                for material_id_str, material in material_objs.items():
                    used_list = self._parse_used_list(material.used_by_products)
                    if product.id not in used_list:
                        used_list.append(product.id)
                        material.used_by_products = self._serialize_used_list(used_list)
                
                self.logger.info(f'产品添加成功: {product.id} - {name}')
                return {'success': True, 'product_id': product.id}
        except Exception as e:
            self.logger.error(f'产品添加异常: {name} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '添加失败'}
    
    def _process_products(self, products):
        """处理配方数据，计算可制作数量"""
        materials_result = self.material_service.get_all_materials()
        materials = {str(m['id']): m for m in materials_result.get('materials', [])}
        
        result = []
        for product in products:
            try:
                product_materials = json.loads(product.materials)
            except:
                product_materials = {}
            
            possible_quantity = float('inf')
            materials_with_stock = []
            
            for material_id, required_qty in product_materials.items():
                material = materials.get(material_id)
                if material:
                    stock_count = material['stock_count']
                    materials_with_stock.append({
                        'product_id': material_id,
                        'name': material['name'],
                        'required': required_qty,
                        'stock_count': stock_count
                    })
                    if required_qty > 0:
                        possible_quantity = min(possible_quantity, stock_count // required_qty)
                else:
                    materials_with_stock.append({
                        'product_id': material_id,
                        'name': material_id,
                        'required': required_qty,
                        'stock_count': 0
                    })
                    possible_quantity = 0
            
            result.append({
                'id': product.id,
                'name': product.name,
                'materials': materials_with_stock,
                'in_price': product.in_price,
                'out_price': product.out_price,
                'other_price': product.other_price or 0,
                'image_path': product.image_path,
                'stock_count': product.stock_count or 0,
                'possible_quantity': int(possible_quantity) if possible_quantity != float('inf') else 0,
                'created_at': format_china_time(product.created_at),
                'updated_at': format_china_time(product.updated_at)
            })
        
        return result
    
    def get_all_products(self):
        """获取所有配方并计算可制作数量"""
        try:
            with self.db.session_scope() as session:
                products = session.query(Product).order_by(Product.id).all()
                return {'success': True, 'products': self._process_products(products)}
        except Exception as e:
            self.logger.error(f'获取所有产品失败: {str(e)}', exc_info=True)
            return {'success': False, 'message': '获取产品列表失败', 'products': []}
    
    def get_products_paginated(self, offset: int, limit: int):
        """分页获取产品"""
        with self.db.session_scope() as session:
            products = session.query(Product).order_by(Product.id).offset(offset).limit(limit).all()
            return {'success': True, 'products': self._process_products(products)}
    
    def get_product_category(self) -> dict:
        """获取产品种类"""
        with self.db.session_scope() as session:
            return {'success': True, 'count': session.query(Product).count()}

    def get_products_count(self) -> dict:
        """获取产品总数"""
        with self.db.session_scope() as session:
            products = session.query(Product).all()
            total_count = sum(product.stock_count or 0 for product in products)
            return {'success': True, 'count': total_count}
    
    def delete_product(self, product_id: int) -> dict:
        """删除产品"""
        try:
            with self.db.session_scope() as session:
                product = session.query(Product).filter(Product.id == product_id).first()
                if not product:
                    return {'success': False, 'message': '产品不存在'}
                
                if product.stock_count > 0:
                    return {'success': False, 'message': f'产品 {product.name} 已制作数量不为零（{product.stock_count}个），请先出库或还原后再删除'}
                
                materials = json.loads(product.materials or '{}')
                if not self._update_material_references(session, product_id, materials, {}):
                    return {'success': False, 'message': '更新材料引用失败'}
                
                session.delete(product)
                self.logger.info(f'产品删除成功: {product_id} - {product.name}')
                return {'success': True}
        except Exception as e:
            self.logger.error(f'产品删除异常: {product_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '删除失败'}
    
    def batch_delete_products(self, product_ids: list) -> dict:
        """批量删除产品"""
        try:
            with self.db.session_scope() as session:
                products = session.query(Product).filter(Product.id.in_(product_ids)).all()
                failed_products = []
                deleted_count = 0
                
                for product in products:
                    if product.stock_count and product.stock_count > 0:
                        failed_products.append({
                            'name': product.name,
                            'reason': f'已制作数量不为零（{product.stock_count}个），请先出库或还原后再删除'
                        })
                        continue
                    
                    materials = json.loads(product.materials or '{}')
                    if not self._update_material_references(session, product.id, materials, {}):
                        failed_products.append({'name': product.name, 'reason': '更新材料引用失败'})
                        continue
                    
                    session.delete(product)
                    deleted_count += 1
                
                self.logger.info(f'批量删除产品: 成功{deleted_count}个')
                if failed_products:
                    return {
                        'success': False,
                        'deleted_count': deleted_count,
                        'failed_products': failed_products,
                        'should_close': deleted_count > 0
                    }
                return {'success': True, 'message': f'成功删除 {deleted_count} 个产品'}
        except Exception as e:
            self.logger.error(f'批量删除产品异常: {str(e)}', exc_info=True)
            return {'success': False, 'message': '删除失败'}

    def update_product(self, product_id: int, name: str, materials: dict = None, in_price: float = None, out_price: float = None, other_price: float = None, image_path: str = None) -> dict:
        """更新产品"""
        try:
            with self.db.session_scope() as session:
                product = session.query(Product).filter(Product.id == product_id).first()
                if not product:
                    return {'success': False, 'message': '产品不存在'}
                
                if materials is not None:
                    valid, error_msg, material_objs = self._validate_materials(session, materials)
                    if not valid:
                        return {'success': False, 'message': error_msg}
                    
                    old_materials = json.loads(product.materials or '{}')
                    if not self._update_material_references(session, product_id, old_materials, materials):
                        return {'success': False, 'message': '更新材料引用失败'}
                    
                    product.materials = json.dumps(materials)
                    
                    if materials:
                        product.in_price = sum((material_objs.get(mid).in_price or 0) * qty for mid, qty in materials.items())
                
                product.name = name
                if other_price is not None:
                    product.other_price = other_price
                if in_price is not None:
                    product.in_price = in_price
                if out_price is not None:
                    product.out_price = out_price
                elif materials and out_price is None:
                    product.out_price = (product.in_price or 0) + (product.other_price or 0)
                if image_path is not None:
                    product.image_path = image_path
                
                self.logger.info(f'产品更新成功: {product_id} - {name}')
                return {'success': True}
        except Exception as e:
            self.logger.error(f'产品更新异常: {product_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '更新失败'}
    
    def inbound(self, product_id: int, quantity: int, customer: str = '') -> dict:
        """入库产品"""
        try:
            with self.db.session_scope() as session:
                product = session.query(Product).filter(Product.id == product_id).first()
                if not product:
                    return {'success': False, 'message': '产品不存在'}
                
                materials = json.loads(product.materials)
                material_ids = [int(mid) for mid in materials.keys()]
                materials_objs = session.query(Material).filter(Material.id.in_(material_ids)).all()
                materials_map = {m.id: m for m in materials_objs}
                
                for material_id_str, required_qty in materials.items():
                    material = materials_map.get(int(material_id_str))
                    if not material or material.stock_count < required_qty * quantity:
                        return {'success': False, 'message': f'材料库存不足: {material.name if material else material_id_str}'}
                
                for material_id_str, required_qty in materials.items():
                    materials_map[int(material_id_str)].stock_count -= required_qty * quantity
                
                product.stock_count = (product.stock_count or 0) + quantity
                self.logger.info(f'产品入库成功: {product_id}, 数量: {quantity}')
                return {'success': True, 'product_name': product.name}
        except Exception as e:
            self.logger.error(f'产品入库异常: {product_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '入库失败'}
    
    def outbound(self, product_id: int, quantity: int, customer: str = '') -> dict:
        """出库产品"""
        try:
            with self.db.session_scope() as session:
                product = session.query(Product).filter(Product.id == product_id).first()
                if not product:
                    return {'success': False, 'message': '产品不存在'}
                
                stock_count = product.stock_count or 0
                if stock_count < quantity:
                    return {'success': False, 'message': f'产品库存不足，当前: {stock_count}'}
                
                product.stock_count = stock_count - quantity
                self.logger.info(f'产品出库成功: {product_id}, 数量: {quantity}')
                return {'success': True, 'product_name': product.name}
        except Exception as e:
            self.logger.error(f'出库产品异常: {product_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '出库失败'}
    
    def restore(self, product_id: int, quantity: int, reason: str = '') -> dict:
        """产品还原"""
        try:
            with self.db.session_scope() as session:
                product = session.query(Product).filter(Product.id == product_id).first()
                if not product or (product.stock_count or 0) < quantity:
                    return {'success': False, 'message': '产品不存在或库存不足'}
                
                materials = json.loads(product.materials)
                material_ids = [int(mid) for mid in materials.keys()]
                materials_objs = session.query(Material).filter(Material.id.in_(material_ids)).all()
                materials_map = {m.id: m for m in materials_objs}
                
                for material_id_str, required_qty in materials.items():
                    material = materials_map.get(int(material_id_str))
                    if material:
                        material.stock_count += required_qty * quantity
                
                product.stock_count = (product.stock_count or 0) - quantity
                self.logger.info(f'产品还原成功: {product_id}, 数量: {quantity}')
                return {'success': True, 'product_name': product.name}
        except Exception as e:
            self.logger.error(f'产品还原异常: {product_id} - {str(e)}', exc_info=True)
            return {'success': False, 'message': '还原失败'}

    def get_stock(self, product_id: int) -> dict:
        """获取产品库存"""
        with self.db.session_scope() as session:
            product = session.query(Product).filter(Product.id == product_id).first()
            if not product:
                return {'success': False, 'message': '产品不存在'}
            return {'success': True, 'stock': product.stock_count or 0}

    def import_from_excel(self, file) -> dict:
        """从Excel导入产品"""
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            failed_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                first_col = str(row[0]) if row[0] else ''
                is_new_format = len(row) >= 6 and (not row[0] or first_col.startswith(('http', 'images/', '#')))
                
                if is_new_format:
                    name_idx, cost_idx, sell_idx, manual_idx, stock_idx = 1, 2, 3, 4, 5
                else:
                    name_idx, cost_idx, sell_idx, manual_idx, stock_idx = 0, 1, 2, 3, 4
                
                if not row[name_idx]:
                    continue
                
                try:
                    name = str(row[name_idx]).strip()
                    in_price = float(row[cost_idx]) if cost_idx < len(row) and row[cost_idx] else 0
                    out_price = float(row[sell_idx]) if sell_idx < len(row) and row[sell_idx] else in_price
                    other_price = float(row[manual_idx]) if manual_idx < len(row) and row[manual_idx] else 0
                    import_stock = int(row[stock_idx]) if stock_idx < len(row) and row[stock_idx] else 0
                    
                    with self.db.session_scope() as session:
                        existing = session.query(Product).filter(Product.name == name).first()
                        
                        if existing:
                            existing.in_price = in_price
                            existing.out_price = out_price
                            existing.other_price = other_price
                            if import_stock > 0:
                                existing.stock_count = (existing.stock_count or 0) + import_stock
                            updated_count += 1
                        else:
                            product = Product(
                                name=name,
                                materials='{}',
                                in_price=in_price,
                                out_price=out_price,
                                other_price=other_price,
                                stock_count=import_stock
                            )
                            session.add(product)
                            created_count += 1
                except Exception as e:
                    self.logger.error(f'第{row_num}行导入失败: {str(e)}')
                    failed_count += 1
            
            msg = f'导入完成，成功 {created_count + updated_count} 个，失败 {failed_count} 个'
            return {
                'success': True,
                'message': msg,
                'total_count': created_count + updated_count + failed_count,
                'created_count': created_count,
                'updated_count': updated_count
            }
        except Exception as e:
            self.logger.error(f'导入Excel失败: {str(e)}', exc_info=True)
            return {'success': False, 'message': f'导入失败: {str(e)}'}
    
    def export_to_excel(self, product_ids: list = None):
        """导出产品到Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '产品列表'
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            
            headers = ['图片', '产品名称', '材料清单', '成本价', '售价', '其它费用', '库存数量', '可制作数量', '创建时间', '更新时间']
            ws.append(headers)
            
            with self.db.session_scope() as session:
                query = session.query(Product).order_by(Product.id)
                if product_ids:
                    query = query.filter(Product.id.in_(product_ids))
                
                products = query.all()
                processed_products = self._process_products(products)
                
                row_idx = 2
                for product in processed_products:
                    materials_text = ''
                    if product.get('materials') and isinstance(product['materials'], list):
                        material_list = [f"{m.get('name', m.get('product_id', ''))}×{m.get('required', 0)}" for m in product['materials']]
                        materials_text = ', '.join(material_list)
                    
                    ws.append([
                        '',
                        product['name'],
                        materials_text,
                        product['in_price'],
                        product['out_price'],
                        product['other_price'],
                        product['stock_count'],
                        product['possible_quantity'],
                        product['created_at'],
                        product['updated_at']
                    ])
                    
                    ws.row_dimensions[row_idx].height = 50
                    
                    if product.get('image_path'):
                        try:
                            image_path = product['image_path']
                            if image_path.startswith(('http://', 'https://')):
                                ws[f'A{row_idx}'] = image_path
                            else:
                                image_file = os.path.join(Config.UPLOAD_FOLDER, os.path.basename(image_path))
                                if os.path.exists(image_file):
                                    img = Image.open(image_file)
                                    img.thumbnail((50, 50))
                                    
                                    img_buffer = BytesIO()
                                    img.save(img_buffer, format='PNG')
                                    img_buffer.seek(0)
                                    
                                    xl_img = XLImage(img_buffer)
                                    xl_img.width = 50
                                    xl_img.height = 50
                                    ws.add_image(xl_img, f'A{row_idx}')
                        except Exception as e:
                            self.logger.warning(f'插入图片失败: {product["name"]} - {str(e)}')
                    
                    ws[f'C{row_idx}'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    row_idx += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            self.logger.error(f'导出Excel失败: {str(e)}', exc_info=True)
            return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
    
    def generate_import_template(self):
        """生成产品导入模板"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '产品导入模板'
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            
            headers = ['产品名称', '成本价', '售价', '其它费用', '库存数量']
            ws.append(headers)
            
            ws.append(['红玛瑙手串', 25.50, 35.00, 5.00, 10])
            ws.append(['蓝宝石项链', 45.00, 65.00, 10.00, 5])
            
            ws['A5'] = '导入说明：'
            ws['A6'] = '1. 产品名称：必填，不能重复，最多100个字符'
            ws['A7'] = '2. 成本价：必填，数字类型'
            ws['A8'] = '3. 售价：必填，数字类型'
            ws['A9'] = '4. 其它费用：可选，默认为0，数字类型'
            ws['A10'] = '5. 库存数量：可选，默认为0，导入后会加到库存中'
            ws['A11'] = '6. 如果产品名称已存在，将更新该产品的信息'
            ws['A12'] = '7. 导入不支持图片，请在导入后手动上传产品图片'
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='产品导入模板.xlsx'
            )
        except Exception as e:
            self.logger.error(f'生成产品导入模板失败: {str(e)}', exc_info=True)
            return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'}), 500
